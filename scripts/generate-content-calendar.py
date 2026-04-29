"""
Generate Q1 2026 content calendar Excel file at
apps/web/public/internal/content-calendar-q1.xlsx.

One-shot script — Prompt 23 PR-B deliverable. Run with:
    py scripts/generate-content-calendar.py
or:
    python scripts/generate-content-calendar.py

The output is gitignored from /sitemap.xml + Disallow'd in /robots.txt under
the /internal/ path. It's a planning artifact, not a public asset.
"""
from __future__ import annotations

from pathlib import Path
from typing import TypedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


class Row(TypedDict):
    working_title: str
    target_keyword: str
    intent: str  # informational | transactional | navigational
    persona: str  # BD-Sponsor | Reg-Affairs | QA-Director | Procurement
    target_page: str
    pillar: str
    word_count: int
    author: str
    status: str  # planned | drafted | in-review | published
    due_date: str  # YYYY-MM-DD


HEADERS = [
    "working_title",
    "target_keyword",
    "intent",
    "persona",
    "target_page",
    "pillar",
    "word_count",
    "author",
    "status",
    "due_date",
]

ROWS: list[Row] = [
    # ---------------- Regulatory pillar (10 rows) ----------------
    {
        "working_title": "GUI-0002 unpacked: a foreign sponsor's reading of the DEL application form",
        "target_keyword": "Health Canada DEL application GUI-0002",
        "intent": "informational",
        "persona": "Reg-Affairs",
        "target_page": "/services/regulatory-services/health-canada-del-licensing",
        "pillar": "regulatory",
        "word_count": 2200,
        "author": "Editorial — regulatory practice",
        "status": "published",
        "due_date": "2026-01-08",
    },
    {
        "working_title": "DEL site change notifications: the pre-implementation form, with examples",
        "target_keyword": "DEL site change notification Canada",
        "intent": "informational",
        "persona": "QA-Director",
        "target_page": "/services/regulatory-services/health-canada-del-licensing",
        "pillar": "regulatory",
        "word_count": 1600,
        "author": "Editorial — regulatory practice",
        "status": "in-review",
        "due_date": "2026-01-22",
    },
    {
        "working_title": "ANDA Module 3 quality data: what FDA actually wants in 3.2.S.4.4",
        "target_keyword": "ANDA Module 3 specification justification",
        "intent": "informational",
        "persona": "Reg-Affairs",
        "target_page": "/services/regulatory-services/us-fda-submissions",
        "pillar": "regulatory",
        "word_count": 1900,
        "author": "Editorial — regulatory practice",
        "status": "in-review",
        "due_date": "2026-02-05",
    },
    {
        "working_title": "DMF Type II reference rights: when a Letter of Authorization isn't enough",
        "target_keyword": "DMF Type II Letter of Authorization",
        "intent": "informational",
        "persona": "Reg-Affairs",
        "target_page": "/services/regulatory-services/us-fda-submissions",
        "pillar": "regulatory",
        "word_count": 1400,
        "author": "Editorial — regulatory practice",
        "status": "drafted",
        "due_date": "2026-02-12",
    },
    {
        "working_title": "eCTD 4.0 transition: regional artifacts that change for Canada and the US",
        "target_keyword": "eCTD 4.0 Canada US transition",
        "intent": "informational",
        "persona": "Reg-Affairs",
        "target_page": "/services/regulatory-services/ctd-ectd-dossier-preparation",
        "pillar": "regulatory",
        "word_count": 1800,
        "author": "Editorial — regulatory practice",
        "status": "in-review",
        "due_date": "2026-02-19",
    },
    {
        "working_title": "Pre-audit gap closure: the four GMP findings we see most often before a Health Canada inspection",
        "target_keyword": "Health Canada GMP inspection findings",
        "intent": "informational",
        "persona": "QA-Director",
        "target_page": "/services/regulatory-services/gmp-audit-preparation",
        "pillar": "regulatory",
        "word_count": 1700,
        "author": "Editorial — regulatory practice",
        "status": "drafted",
        "due_date": "2026-02-26",
    },
    {
        "working_title": "USFDA establishment inspection 483 patterns 2024-2025 and what they mean for sterile sites",
        "target_keyword": "FDA Form 483 sterile inspection 2024",
        "intent": "informational",
        "persona": "QA-Director",
        "target_page": "/services/regulatory-services/gmp-audit-preparation",
        "pillar": "regulatory",
        "word_count": 2000,
        "author": "Editorial — regulatory practice",
        "status": "in-review",
        "due_date": "2026-03-05",
    },
    {
        "working_title": "Lifecycle management of an ANDA: change-being-effected vs prior-approval supplements, with thresholds",
        "target_keyword": "ANDA CBE prior approval supplement",
        "intent": "informational",
        "persona": "Reg-Affairs",
        "target_page": "/services/regulatory-services/lifecycle-management",
        "pillar": "regulatory",
        "word_count": 1500,
        "author": "Editorial — regulatory practice",
        "status": "in-review",
        "due_date": "2026-03-12",
    },
    {
        "working_title": "Health Canada NOC/c lifecycle obligations and the post-market commitment record",
        "target_keyword": "Health Canada NOC c post-market commitments",
        "intent": "informational",
        "persona": "Reg-Affairs",
        "target_page": "/services/regulatory-services/lifecycle-management",
        "pillar": "regulatory",
        "word_count": 1600,
        "author": "Editorial — regulatory practice",
        "status": "planned",
        "due_date": "2026-03-19",
    },
    {
        "working_title": "Foreign sponsor's first DEL: scoping the 250-day Health Canada service standard against your launch plan",
        "target_keyword": "Health Canada DEL service standard 250 days",
        "intent": "transactional",
        "persona": "BD-Sponsor",
        "target_page": "/services/regulatory-services/health-canada-del-licensing",
        "pillar": "regulatory",
        "word_count": 2400,
        "author": "Editorial — regulatory practice",
        "status": "planned",
        "due_date": "2026-03-26",
    },
    # ---------------- Drug-development pillar (8 rows) ----------------
    {
        "working_title": "BCS Class II oral solids: dissolution method choices that survive ICH Q14 review",
        "target_keyword": "BCS class II dissolution ICH Q14",
        "intent": "informational",
        "persona": "Reg-Affairs",
        "target_page": "/services/pharmaceutical-development/solid-oral-dosage",
        "pillar": "drug-development",
        "word_count": 1800,
        "author": "Editorial — analytical bench",
        "status": "published",
        "due_date": "2026-01-15",
    },
    {
        "working_title": "Pediatric oral liquids: taste-masking, preservative selection, and the stability data set FDA expects",
        "target_keyword": "pediatric oral liquid formulation taste masking",
        "intent": "informational",
        "persona": "BD-Sponsor",
        "target_page": "/services/pharmaceutical-development/liquid-oral-dosage",
        "pillar": "drug-development",
        "word_count": 1900,
        "author": "Editorial — analytical bench",
        "status": "drafted",
        "due_date": "2026-01-29",
    },
    {
        "working_title": "Sterile injectable container closure integrity: dye ingress, vacuum decay, or HSA — with one decision tree",
        "target_keyword": "container closure integrity testing CCIT methods",
        "intent": "informational",
        "persona": "QA-Director",
        "target_page": "/services/pharmaceutical-development/sterile-injectables",
        "pillar": "drug-development",
        "word_count": 1700,
        "author": "Editorial — analytical bench",
        "status": "in-review",
        "due_date": "2026-02-05",
    },
    {
        "working_title": "Modified-release oral solids: f1/f2 dissolution similarity in a Q1/Q2 generic context",
        "target_keyword": "f2 similarity factor modified release",
        "intent": "informational",
        "persona": "Reg-Affairs",
        "target_page": "/services/pharmaceutical-development/solid-oral-dosage",
        "pillar": "drug-development",
        "word_count": 1500,
        "author": "Editorial — analytical bench",
        "status": "drafted",
        "due_date": "2026-02-12",
    },
    {
        "working_title": "Topical semisolid Q3 microstructure characterization for ANDA filings",
        "target_keyword": "topical semisolid Q3 microstructure ANDA",
        "intent": "informational",
        "persona": "Reg-Affairs",
        "target_page": "/services/pharmaceutical-development/topical-semisolid",
        "pillar": "drug-development",
        "word_count": 1800,
        "author": "Editorial — analytical bench",
        "status": "drafted",
        "due_date": "2026-02-19",
    },
    {
        "working_title": "MDI/DPI device-formulation interaction studies: what to add to your stability protocol",
        "target_keyword": "MDI DPI device formulation interaction stability",
        "intent": "informational",
        "persona": "QA-Director",
        "target_page": "/services/pharmaceutical-development/inhalation",
        "pillar": "drug-development",
        "word_count": 2000,
        "author": "Editorial — analytical bench",
        "status": "planned",
        "due_date": "2026-03-05",
    },
    {
        "working_title": "Transdermal patch adhesion testing: ASTM, USP, and the comparator-product data sponsors miss",
        "target_keyword": "transdermal patch adhesion ASTM USP",
        "intent": "informational",
        "persona": "Reg-Affairs",
        "target_page": "/services/pharmaceutical-development/transdermal-modified-release",
        "pillar": "drug-development",
        "word_count": 1500,
        "author": "Editorial — analytical bench",
        "status": "planned",
        "due_date": "2026-03-12",
    },
    {
        "working_title": "Tech transfer to Mississauga: the data the receiving site needs in week one",
        "target_keyword": "tech transfer receiving site week one",
        "intent": "transactional",
        "persona": "BD-Sponsor",
        "target_page": "/services/pharmaceutical-development/solid-oral-dosage",
        "pillar": "drug-development",
        "word_count": 1300,
        "author": "Editorial — operating model",
        "status": "drafted",
        "due_date": "2026-03-19",
    },
    # ---------------- Analytical pillar (6 rows) ----------------
    {
        "working_title": "ICH Q2(R2) and ICH Q14: what the analytical procedure life-cycle looks like in 2026",
        "target_keyword": "ICH Q2 R2 Q14 analytical lifecycle",
        "intent": "informational",
        "persona": "Reg-Affairs",
        "target_page": "/services/analytical-services/method-validation",
        "pillar": "analytical",
        "word_count": 2100,
        "author": "Editorial — analytical bench",
        "status": "published",
        "due_date": "2026-01-08",
    },
    {
        "working_title": "Stability program design under ICH Q1A(R2): pulling protocol decisions back into formulation",
        "target_keyword": "ICH Q1A stability protocol design",
        "intent": "informational",
        "persona": "QA-Director",
        "target_page": "/services/analytical-services/stability-studies",
        "pillar": "analytical",
        "word_count": 1700,
        "author": "Editorial — analytical bench",
        "status": "in-review",
        "due_date": "2026-01-22",
    },
    {
        "working_title": "Genotoxic impurities under ICH M7(R2): permitted daily exposure, control strategy, and a worked example",
        "target_keyword": "ICH M7 R2 genotoxic impurity PDE",
        "intent": "informational",
        "persona": "Reg-Affairs",
        "target_page": "/services/analytical-services/impurity-profiling",
        "pillar": "analytical",
        "word_count": 1800,
        "author": "Editorial — analytical bench",
        "status": "drafted",
        "due_date": "2026-02-05",
    },
    {
        "working_title": "LC-MS/MS bioanalytical method validation: incurred sample reanalysis the way M10 actually reads",
        "target_keyword": "ICH M10 bioanalytical incurred sample reanalysis",
        "intent": "informational",
        "persona": "Reg-Affairs",
        "target_page": "/services/analytical-services/bioanalytical",
        "pillar": "analytical",
        "word_count": 1600,
        "author": "Editorial — analytical bench",
        "status": "drafted",
        "due_date": "2026-02-19",
    },
    {
        "working_title": "Extractables and leachables for OTC injectables: USP <1663> and <1664> read together",
        "target_keyword": "extractables leachables USP 1663 1664",
        "intent": "informational",
        "persona": "QA-Director",
        "target_page": "/services/analytical-services/extractables-and-leachables",
        "pillar": "analytical",
        "word_count": 1900,
        "author": "Editorial — analytical bench",
        "status": "planned",
        "due_date": "2026-03-05",
    },
    {
        "working_title": "Reference standard qualification under USP <11>: when to characterize in-house vs source from USP",
        "target_keyword": "USP reference standard characterization 11",
        "intent": "informational",
        "persona": "QA-Director",
        "target_page": "/services/analytical-services/reference-standard-characterization",
        "pillar": "analytical",
        "word_count": 1400,
        "author": "Editorial — analytical bench",
        "status": "planned",
        "due_date": "2026-03-19",
    },
    # ---------------- Quality pillar (3 rows) ----------------
    {
        "working_title": "CAPA effectiveness checks Health Canada inspectors actually open: writing the verification plan",
        "target_keyword": "CAPA effectiveness verification Health Canada",
        "intent": "informational",
        "persona": "QA-Director",
        "target_page": "/quality-compliance",
        "pillar": "quality",
        "word_count": 1500,
        "author": "Editorial — operating model",
        "status": "in-review",
        "due_date": "2026-01-29",
    },
    {
        "working_title": "From documents to data: the QMS controls a 2026 GMP audit expects you to evidence digitally",
        "target_keyword": "digital QMS GMP audit evidence 2026",
        "intent": "informational",
        "persona": "QA-Director",
        "target_page": "/quality-compliance",
        "pillar": "quality",
        "word_count": 1700,
        "author": "Editorial — operating model",
        "status": "drafted",
        "due_date": "2026-02-26",
    },
    {
        "working_title": "Supplier qualification under ICH Q10: the audit cadence, the documentation set, and where most files thin out",
        "target_keyword": "ICH Q10 supplier qualification audit cadence",
        "intent": "informational",
        "persona": "Procurement",
        "target_page": "/quality-compliance",
        "pillar": "quality",
        "word_count": 1600,
        "author": "Editorial — operating model",
        "status": "planned",
        "due_date": "2026-03-26",
    },
    # ---------------- Manufacturing pillar (2 rows) ----------------
    {
        "working_title": "Process validation Stage 2 PPQ for oral solids: batch count, sampling, and the protocol that survives review",
        "target_keyword": "process performance qualification PPQ oral solid",
        "intent": "informational",
        "persona": "QA-Director",
        "target_page": "/services/pharmaceutical-development/solid-oral-dosage",
        "pillar": "manufacturing",
        "word_count": 1800,
        "author": "Editorial — operating model",
        "status": "drafted",
        "due_date": "2026-02-12",
    },
    {
        "working_title": "Scale-up factor for high-shear granulation: an honest look at where the model breaks",
        "target_keyword": "high shear granulation scale up factor",
        "intent": "informational",
        "persona": "BD-Sponsor",
        "target_page": "/services/pharmaceutical-development/solid-oral-dosage",
        "pillar": "manufacturing",
        "word_count": 1400,
        "author": "Editorial — operating model",
        "status": "published",
        "due_date": "2026-03-12",
    },
    # ---------------- Logistics pillar (1 row) ----------------
    {
        "working_title": "3PL distribution under a Health Canada DEL: temperature mapping, pedigree, and the importer-of-record question",
        "target_keyword": "3PL Canada DEL temperature mapping importer of record",
        "intent": "transactional",
        "persona": "Procurement",
        "target_page": "/facilities/mississauga-canada",
        "pillar": "logistics",
        "word_count": 1500,
        "author": "Editorial — operating model",
        "status": "published",
        "due_date": "2026-02-26",
    },
]

# Sanity checks (run at script time, fail loud).
assert len(ROWS) == 30, f"Expected 30 rows, got {len(ROWS)}"
PILLAR_COUNTS = {
    "regulatory": 10,
    "drug-development": 8,
    "analytical": 6,
    "quality": 3,
    "manufacturing": 2,
    "logistics": 1,
}
actual_pillar_counts: dict[str, int] = {}
for row in ROWS:
    actual_pillar_counts[row["pillar"]] = (
        actual_pillar_counts.get(row["pillar"], 0) + 1
    )
for pillar, expected in PILLAR_COUNTS.items():
    assert actual_pillar_counts.get(pillar) == expected, (
        f"Pillar {pillar}: expected {expected}, got {actual_pillar_counts.get(pillar, 0)}"
    )
STATUS_COUNTS = {"published": 5, "in-review": 8, "drafted": 10, "planned": 7}
actual_status_counts: dict[str, int] = {}
for row in ROWS:
    actual_status_counts[row["status"]] = (
        actual_status_counts.get(row["status"], 0) + 1
    )
for status, expected in STATUS_COUNTS.items():
    assert actual_status_counts.get(status) == expected, (
        f"Status {status}: expected {expected}, got {actual_status_counts.get(status, 0)}"
    )

OUT_PATH = (
    Path(__file__).resolve().parent.parent
    / "apps"
    / "web"
    / "public"
    / "internal"
    / "content-calendar-q1.xlsx"
)


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Q1 2026"

    header_fill = PatternFill("solid", start_color="0E4C5A")
    header_font = Font(name="Arial", bold=True, color="FAFAF7", size=11)
    body_font = Font(name="Arial", size=11)

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row in ROWS:
        ws.append([row[h] for h in HEADERS])

    for row in ws.iter_rows(min_row=2, max_row=len(ROWS) + 1):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    table_ref = f"A1:{get_column_letter(len(HEADERS))}{len(ROWS) + 1}"
    table = Table(displayName="ContentCalendarQ1", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    column_widths = {
        "working_title": 64,
        "target_keyword": 38,
        "intent": 14,
        "persona": 14,
        "target_page": 56,
        "pillar": 18,
        "word_count": 12,
        "author": 32,
        "status": 12,
        "due_date": 14,
    }
    for idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = column_widths[header]

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    return wb


def main() -> None:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH} — {len(ROWS)} rows")


if __name__ == "__main__":
    main()
