"""
Generate the 301 redirect map TEMPLATE as docs/redirects-301-map.xlsx.

Output: docs/redirects-301-map.xlsx — a structured template that ops
fills in with legacy-URL ↔ new-URL pairs before launch. The template
ships with the new-IA target column pre-populated to the public route
inventory and four worked examples (the existing redirects already in
apps/web/next.config.ts).

This is a one-shot script. Re-run when the new-IA route inventory in
PAGES below changes — but DO NOT regenerate the file once ops has
started filling in the legacy column. To detect that, the script
compares against an existing file: if the existing file has any
non-template legacy-path values, refuse to overwrite without an
explicit --force flag.

Run:
    py scripts/generate-301-redirect-map.py
    py scripts/generate-301-redirect-map.py --force   # overwrite template

Requirements:
    py -m pip install openpyxl==3.1.2

Mirrors scripts/generate-qa-matrix.py on style. Brand palette
duplicated from packages/config/design-tokens.css.
"""

from __future__ import annotations

import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    print(
        "Missing dependency. Install with:\n  py -m pip install openpyxl==3.1.2",
        file=sys.stderr,
    )
    raise SystemExit(1) from exc


BRAND_PRIMARY_HEX = "0E4C5A"
BRAND_FG_HEX = "1F2933"
BRAND_TABLE_HEADER_BG = "EBF4F6"
PLACEHOLDER_FILL = "FFF7E6"  # warm-cream highlight for "fill me in" cells


REPO_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_PATH = REPO_ROOT / "docs" / "redirects-301-map.xlsx"


# ---------------------------------------------------------------------------
# Source data
# ---------------------------------------------------------------------------


# Worked examples — the four redirects already shipping in
# apps/web/next.config.ts. These are committed-to-prod truths, not
# placeholders, so ops sees a real example of the schema.
WORKED_EXAMPLES: list[dict[str, str]] = [
    {
        "legacy_path": "/whitepapers/canada-india-playbook",
        "new_path": "/insights/whitepapers/canadian-cdmo-operating-model",
        "redirect_type": "301",
        "rationale": "Prompt 15 rebrand — wired in next.config.ts",
        "owner": "Engineering",
        "status": "Deployed",
    },
    {
        "legacy_path": "/insights/whitepapers/two-hub-operating-model",
        "new_path": "/insights/whitepapers/canadian-cdmo-operating-model",
        "redirect_type": "301",
        "rationale": "Canadian rebrand — wired in next.config.ts",
        "owner": "Engineering",
        "status": "Deployed",
    },
    {
        "legacy_path": "/insights/inside-a-two-hub-cdmo",
        "new_path": "/insights/inside-our-operating-model",
        "redirect_type": "301",
        "rationale": "Article slug retired — wired in next.config.ts",
        "owner": "Engineering",
        "status": "Deployed",
    },
    {
        "legacy_path": "/downloads/two-hub-operating-model.pdf",
        "new_path": "/downloads/canadian-cdmo-operating-model.pdf",
        "redirect_type": "301",
        "rationale": "PDF asset path — wired in next.config.ts",
        "owner": "Engineering",
        "status": "Deployed",
    },
]


# Suggested target routes — the publicly-resolvable canonical IA.
# Used to populate the second sheet so ops can see what URLs the
# legacy site SHOULD map to.
TARGET_ROUTES: list[tuple[str, str]] = [
    ("/", "Home"),
    ("/why-propharmex", "Why Propharmex"),
    ("/about", "About"),
    ("/about/leadership", "Leadership"),
    ("/quality-compliance", "Quality + compliance"),
    ("/facilities", "Facilities hub"),
    ("/facilities/mississauga-canada", "Mississauga site"),
    ("/facilities/hyderabad-india", "Hyderabad development centre"),
    ("/services", "Services hub"),
    ("/services/pharmaceutical-development/<dosage-form>", "Pharmaceutical development leaves"),
    ("/services/analytical-services/<service>", "Analytical services leaves"),
    ("/services/regulatory-services/<service>", "Regulatory services leaves"),
    ("/industries", "Industries hub"),
    ("/industries/<slug>", "Industry detail"),
    ("/case-studies", "Case studies hub"),
    ("/case-studies/<slug>", "Case study detail"),
    ("/insights", "Insights hub"),
    ("/insights/<slug>", "Article detail"),
    ("/insights/whitepapers/<slug>", "Whitepaper detail"),
    ("/our-process", "Our process"),
    ("/contact", "Contact"),
    ("/ai/project-scoping-assistant", "AI tool — Scoping"),
    ("/ai/del-readiness", "AI tool — DEL Readiness"),
    ("/ai/dosage-matcher", "AI tool — Dosage Matcher"),
    ("/accessibility", "Accessibility statement"),
]


# Placeholder rows ops fills in. Pre-populated with TBD markers so
# the template is obviously incomplete on first open.
TEMPLATE_ROW_COUNT = 30  # adjust as legacy URL count is known


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------


def hex_fill(hex_str: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_str)


def thin_border() -> Border:
    side = Side(style="thin", color="D6D8DD")
    return Border(top=side, left=side, right=side, bottom=side)


def style_header_cell(cell) -> None:
    cell.font = Font(name="Calibri", size=10, bold=True, color=BRAND_PRIMARY_HEX)
    cell.fill = hex_fill(BRAND_TABLE_HEADER_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = thin_border()


def style_body_cell(cell, placeholder: bool = False) -> None:
    cell.font = Font(name="Calibri", size=10, color=BRAND_FG_HEX)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = thin_border()
    if placeholder:
        cell.fill = hex_fill(PLACEHOLDER_FILL)


def autosize_columns(ws, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


REDIRECT_HEADERS = [
    "legacy_path",
    "new_path",
    "redirect_type",
    "rationale",
    "owner",
    "status",
]


def build_redirects_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Redirects"

    # Header
    for c_idx, value in enumerate(REDIRECT_HEADERS, start=1):
        cell = ws.cell(row=1, column=c_idx, value=value)
        style_header_cell(cell)
    ws.row_dimensions[1].height = 30

    r_idx = 2

    # Worked examples first (flagged as Deployed)
    for example in WORKED_EXAMPLES:
        for c_idx, key in enumerate(REDIRECT_HEADERS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=example[key])
            style_body_cell(cell)
        r_idx += 1

    # Placeholder rows for ops to fill in
    for _ in range(TEMPLATE_ROW_COUNT):
        ws.cell(row=r_idx, column=1, value="<paste legacy URL path>")
        ws.cell(row=r_idx, column=2, value="<paste new IA path>")
        ws.cell(row=r_idx, column=3, value="301")
        ws.cell(row=r_idx, column=4, value="<one-line rationale>")
        ws.cell(row=r_idx, column=5, value="Operations")
        ws.cell(row=r_idx, column=6, value="Pending")
        for c_idx in range(1, len(REDIRECT_HEADERS) + 1):
            style_body_cell(ws.cell(row=r_idx, column=c_idx), placeholder=True)
        r_idx += 1

    autosize_columns(
        ws,
        {1: 48, 2: 48, 3: 16, 4: 42, 5: 18, 6: 16},
    )
    ws.freeze_panes = "A2"


def build_target_routes_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Target IA")

    headers = ["new_path", "purpose"]
    for c_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=value)
        style_header_cell(cell)
    ws.row_dimensions[1].height = 28

    for r_idx, (path, purpose) in enumerate(TARGET_ROUTES, start=2):
        ws.cell(row=r_idx, column=1, value=path)
        ws.cell(row=r_idx, column=2, value=purpose)
        for c_idx in range(1, len(headers) + 1):
            style_body_cell(ws.cell(row=r_idx, column=c_idx))

    autosize_columns(ws, {1: 56, 2: 56})
    ws.freeze_panes = "A2"


def build_legend_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Legend")

    rows = [
        ("Concept", "Convention"),
        ("Status values", "Pending / Approved / Deployed / Rejected"),
        ("Redirect type", "301 (permanent) for retired URLs; 302 only for time-bounded promotions"),
        ("Owner", "Operations is the default owner for legacy-path mapping; Engineering deploys"),
        (
            "Source of truth",
            "next.config.ts redirects() is the load-bearing source of truth; this xlsx is the planning artifact",
        ),
        (
            "How to deploy",
            "Once a row is Approved by ops, engineering adds it to apps/web/next.config.ts and updates status to Deployed",
        ),
    ]

    for c_idx, value in enumerate(rows[0], start=1):
        cell = ws.cell(row=1, column=c_idx, value=value)
        style_header_cell(cell)
    ws.row_dimensions[1].height = 26

    for r_idx, row in enumerate(rows[1:], start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            style_body_cell(cell)

    autosize_columns(ws, {1: 22, 2: 80})


# ---------------------------------------------------------------------------
# Safety check + main
# ---------------------------------------------------------------------------


def file_has_real_data(path: Path) -> bool:
    """Return True if the existing redirect map has any non-template rows."""
    if not path.exists():
        return False
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Redirects" not in wb.sheetnames:
        return False
    ws = wb["Redirects"]
    deployed_legacy_paths = {ex["legacy_path"] for ex in WORKED_EXAMPLES}
    placeholder_markers = {
        "<paste legacy URL path>",
        "<paste new IA path>",
        "<one-line rationale>",
        "",
        None,
    }
    # Any row whose legacy_path (col 1) is non-empty, not a placeholder, and
    # not one of the four worked examples means ops has filled in real data.
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        legacy = row[0]
        if legacy in placeholder_markers:
            continue
        if legacy in deployed_legacy_paths:
            continue
        return True
    return False


def assert_structure() -> None:
    assert len(WORKED_EXAMPLES) == 4, (
        f"Expected exactly 4 worked examples (the redirects shipping in "
        f"apps/web/next.config.ts), got {len(WORKED_EXAMPLES)}."
    )
    assert len(TARGET_ROUTES) >= 20, (
        f"Expected at least 20 canonical IA routes in TARGET_ROUTES, "
        f"got {len(TARGET_ROUTES)}."
    )
    assert TEMPLATE_ROW_COUNT >= 20, (
        f"Template should leave at least 20 placeholder rows, "
        f"got {TEMPLATE_ROW_COUNT}."
    )


def main() -> None:
    assert_structure()

    force = "--force" in sys.argv
    if file_has_real_data(OUTPUT_PATH) and not force:
        print(
            f"{OUTPUT_PATH.relative_to(REPO_ROOT)} appears to already contain "
            "operations-filled rows (non-template legacy paths beyond the four "
            "worked examples).\n"
            "Refusing to overwrite. Re-run with --force only if you intend to "
            "discard those rows.",
            file=sys.stderr,
        )
        raise SystemExit(2)

    wb = Workbook()
    build_redirects_sheet(wb)
    build_target_routes_sheet(wb)
    build_legend_sheet(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    rel = OUTPUT_PATH.relative_to(REPO_ROOT)
    print(
        f"Wrote {rel} - {len(WORKED_EXAMPLES)} worked examples, "
        f"{TEMPLATE_ROW_COUNT} placeholder rows, "
        f"{len(TARGET_ROUTES)} target routes."
    )


if __name__ == "__main__":
    main()
