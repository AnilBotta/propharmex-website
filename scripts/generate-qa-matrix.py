"""
Generate the launch QA matrix as docs/qa-matrix.xlsx.

Output: docs/qa-matrix.xlsx — the QA team's pre-launch sign-off
template. One sheet per concern: a status grid (page x breakpoint x
browser), a per-page test plan, and a sign-off summary.

This is a one-shot script. The PAGES, BREAKPOINTS, and BROWSERS
constants below are the authoritative source. Edit them in this file
and re-run the script when the page inventory or browser support
target changes.

Run:
    py scripts/generate-qa-matrix.py

Requirements:
    py -m pip install openpyxl==3.1.2

Mirrors scripts/generate-content-calendar.py (Prompt 23) on style and
the .docx generators on brand palette.
"""

from __future__ import annotations

from pathlib import Path
from typing import TypedDict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:
    import sys

    print(
        "Missing dependency. Install with:\n  py -m pip install openpyxl==3.1.2",
        file=sys.stderr,
    )
    raise SystemExit(1) from exc


# Brand tokens duplicated from packages/config/design-tokens.css. Keep
# in sync if the brand palette ever shifts.
BRAND_PRIMARY_HEX = "0E4C5A"
BRAND_FG_HEX = "1F2933"
BRAND_MUTED_HEX = "5F6573"
BRAND_TABLE_HEADER_BG = "EBF4F6"


REPO_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_PATH = REPO_ROOT / "docs" / "qa-matrix.xlsx"


# ---------------------------------------------------------------------------
# Source data — page inventory mirrors docs/handoff.md §4.
# ---------------------------------------------------------------------------


class PageEntry(TypedDict):
    route: str
    purpose: str
    notes: str


PAGES: list[PageEntry] = [
    {"route": "/", "purpose": "Home", "notes": "Region-aware hero ordering"},
    {"route": "/why-propharmex", "purpose": "Marketing — anti-hype proof points", "notes": ""},
    {"route": "/about", "purpose": "Marketing — Canadian anchor narrative", "notes": ""},
    {"route": "/about/leadership", "purpose": "Marketing — leadership grid", "notes": ""},
    {"route": "/quality-compliance", "purpose": "Marketing — region-aware certs", "notes": ""},
    {"route": "/facilities", "purpose": "Marketing — facility hub", "notes": "Mississauga + Hyderabad"},
    {"route": "/facilities/mississauga-canada", "purpose": "Detail — DEL anchor site", "notes": ""},
    {"route": "/facilities/hyderabad-india", "purpose": "Detail — development centre", "notes": ""},
    {"route": "/services", "purpose": "Hub — service tree root", "notes": ""},
    {"route": "/services/pharmaceutical-development/solid-oral-dosage", "purpose": "Service detail — solid oral", "notes": "Sample of the dosage-form set"},
    {"route": "/services/analytical-services/method-development", "purpose": "Service detail — analytical", "notes": "Sample of the analytical-service set"},
    {"route": "/services/regulatory-services/health-canada-del-licensing", "purpose": "Service detail — DEL licensing", "notes": ""},
    {"route": "/industries", "purpose": "Hub — industries", "notes": ""},
    {"route": "/industries/pharmaceutical-innovators", "purpose": "Industry detail", "notes": "Sample"},
    {"route": "/case-studies", "purpose": "Hub — case studies", "notes": ""},
    {"route": "/case-studies/sterile-injectable-second-sourcing", "purpose": "Case study detail", "notes": "Sample"},
    {"route": "/insights", "purpose": "Hub — insights", "notes": "Articles + whitepapers"},
    {"route": "/insights/del-at-a-glance-foreign-sponsor-primer", "purpose": "Article detail", "notes": "Sample"},
    {"route": "/insights/whitepapers/canadian-cdmo-operating-model", "purpose": "Whitepaper gate", "notes": "Resend + Turnstile"},
    {"route": "/our-process", "purpose": "Marketing — engagement workflow", "notes": ""},
    {"route": "/contact", "purpose": "Form — inquiry + Cal.com + addresses", "notes": "Turnstile-gated submit"},
    {"route": "/ai/project-scoping-assistant", "purpose": "AI tool — conversational scope", "notes": "Lazy-imported surface"},
    {"route": "/ai/del-readiness", "purpose": "AI tool — gap assessment", "notes": "PDF download"},
    {"route": "/ai/dosage-matcher", "purpose": "AI tool — formulation match", "notes": "PDF download"},
    {"route": "(all)", "purpose": "Concierge bubble — site-wide", "notes": "Open + ask + cite + close"},
    {"route": "/accessibility", "purpose": "Statement", "notes": ""},
]


BREAKPOINTS: list[tuple[str, int]] = [
    ("Mobile (375)", 375),
    ("Tablet (768)", 768),
    ("Laptop (1280)", 1280),
    ("Desktop (1440)", 1440),
]


BROWSERS: list[str] = ["Chromium", "WebKit", "Firefox"]


# Per-page checklist applied during QA — kept short on purpose.
PAGE_CHECKLIST: list[str] = [
    "Layout intact at the breakpoint (no overflow, no clipped text)",
    "All images load + have meaningful alt text where non-decorative",
    "Tab order follows visible order; visible focus ring on every interactive",
    "No console errors during initial paint or after interaction",
    "Primary CTA navigates to its declared destination",
    "Page passes the in-CI axe scan (zero serious or critical violations)",
    "Page meets the in-CI Lighthouse perf budget (warn at 0.90)",
    "JSON-LD validates in Schema.org Validator",
    "OG tags render correctly when previewed in social-link debugger",
    "Region cookie (px-region) does not break the page",
]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------


def hex_fill(hex_str: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_str)


def thin_border() -> Border:
    side = Side(style="thin", color="D6D8DD")
    return Border(top=side, left=side, right=side, bottom=side)


def style_header_cell(cell) -> None:
    cell.font = Font(name="Calibri", size=10, bold=True, color=BRAND_PRIMARY_HEX)
    cell.fill = hex_fill(BRAND_TABLE_HEADER_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = thin_border()


def style_body_cell(cell) -> None:
    cell.font = Font(name="Calibri", size=10, color=BRAND_FG_HEX)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = thin_border()


def autosize_columns(ws, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def build_status_sheet(wb: Workbook) -> None:
    """Sheet 1: page x breakpoint x browser status grid."""
    ws = wb.active
    ws.title = "Status grid"

    headers = ["Route", "Purpose"]
    for bp_label, _ in BREAKPOINTS:
        for browser in BROWSERS:
            headers.append(f"{bp_label} · {browser}")
    headers.append("Notes")

    # Header row
    for c_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=value)
        style_header_cell(cell)
    ws.row_dimensions[1].height = 38

    # Data rows
    for r_idx, page in enumerate(PAGES, start=2):
        ws.cell(row=r_idx, column=1, value=page["route"])
        ws.cell(row=r_idx, column=2, value=page["purpose"])
        # status cells start at column 3, end at column 2 + (4 BP * 3 BR) = 14
        for c_idx in range(3, 3 + len(BREAKPOINTS) * len(BROWSERS)):
            ws.cell(row=r_idx, column=c_idx, value="")
        ws.cell(
            row=r_idx,
            column=3 + len(BREAKPOINTS) * len(BROWSERS),
            value=page["notes"],
        )
        for c_idx in range(1, len(headers) + 1):
            style_body_cell(ws.cell(row=r_idx, column=c_idx))

    # Column widths: route + purpose wider, status cells narrow, notes wide
    widths = {1: 42, 2: 36}
    for c_idx in range(3, 3 + len(BREAKPOINTS) * len(BROWSERS)):
        widths[c_idx] = 14
    widths[len(headers)] = 38
    autosize_columns(ws, widths)

    ws.freeze_panes = "C2"


def build_test_plan_sheet(wb: Workbook) -> None:
    """Sheet 2: per-page checklist QA team works through."""
    ws = wb.create_sheet("Test plan")

    headers = ["Route", "Purpose", "Checklist item", "Pass/Fail", "Tester", "Notes"]
    for c_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=value)
        style_header_cell(cell)
    ws.row_dimensions[1].height = 32

    r_idx = 2
    for page in PAGES:
        for item in PAGE_CHECKLIST:
            ws.cell(row=r_idx, column=1, value=page["route"])
            ws.cell(row=r_idx, column=2, value=page["purpose"])
            ws.cell(row=r_idx, column=3, value=item)
            ws.cell(row=r_idx, column=4, value="")
            ws.cell(row=r_idx, column=5, value="")
            ws.cell(row=r_idx, column=6, value="")
            for c_idx in range(1, len(headers) + 1):
                style_body_cell(ws.cell(row=r_idx, column=c_idx))
            r_idx += 1

    autosize_columns(ws, {1: 42, 2: 36, 3: 60, 4: 14, 5: 18, 6: 32})
    ws.freeze_panes = "D2"


def build_signoff_sheet(wb: Workbook) -> None:
    """Sheet 3: launch sign-off summary."""
    ws = wb.create_sheet("Sign-off")

    rows = [
        ("Concern", "Owner", "Status", "Date", "Notes"),
        ("Status grid complete (no Sev 1 reds)", "", "", "", ""),
        ("Test plan items all completed or waived", "", "", "", ""),
        ("Lighthouse CI green at 0.90 perf / strict CWV", "", "", "", ""),
        ("Bundle budget (450 kB) green per .github/workflows/bundle-budget.yml", "", "", "", ""),
        ("axe-core CI zero serious or critical", "", "", "", ""),
        ("Manual VoiceOver pass (per docs/accessibility-at-test-plan.md)", "", "", "", ""),
        ("Manual NVDA pass (per docs/accessibility-at-test-plan.md)", "", "", "", ""),
        ("Sentry production source maps uploaded", "", "", "", ""),
        ("PostHog dashboards live (per docs/analytics-taxonomy.md §6)", "", "", "", ""),
        ("Resend production sender domain verified", "", "", "", ""),
        ("Cal.com production link configured (CAL_LINK env var)", "", "", "", ""),
        ("Turnstile production keys deployed", "", "", "", ""),
        ("DNS cutover complete + verified", "", "", "", ""),
        ("301 redirects deployed (per docs/redirects-301-map.xlsx)", "", "", "", ""),
        ("Sitemap submitted to Google Search Console", "", "", "", ""),
        ("Sitemap submitted to Bing Webmaster Tools", "", "", "", ""),
        ("v1.0.0 git tag pushed", "", "", "", ""),
    ]

    for c_idx, value in enumerate(rows[0], start=1):
        cell = ws.cell(row=1, column=c_idx, value=value)
        style_header_cell(cell)
    ws.row_dimensions[1].height = 28

    for r_idx, row in enumerate(rows[1:], start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            style_body_cell(cell)

    autosize_columns(ws, {1: 56, 2: 22, 3: 14, 4: 14, 5: 36})
    ws.freeze_panes = "B2"


def build_legend_sheet(wb: Workbook) -> None:
    """Sheet 4: short notes on conventions used in the matrix."""
    ws = wb.create_sheet("Legend")

    rows = [
        ("Concept", "Convention"),
        ("Status cell values", "Pass / Fail / N/A / Blocked / Skip"),
        ("Severity classes", "Sev 1 = launch blocker; Sev 2 = ship-with-fix-planned; Sev 3 = polish"),
        ("Browsers", "Chromium / WebKit / Firefox via Playwright + manual spot-check"),
        ("Breakpoints", "375 mobile / 768 tablet / 1280 laptop / 1440 desktop"),
        ("Smoke suite", "apps/web/e2e/*.spec.ts — runs in CI; covers happy paths only"),
        ("Out of scope here", "Form-submit correctness (Turnstile-gated); covered by integration tests"),
    ]

    for c_idx, value in enumerate(rows[0], start=1):
        cell = ws.cell(row=1, column=c_idx, value=value)
        style_header_cell(cell)
    ws.row_dimensions[1].height = 24

    for r_idx, row in enumerate(rows[1:], start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            style_body_cell(cell)

    autosize_columns(ws, {1: 28, 2: 80})


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def assert_structure() -> None:
    assert len(PAGES) >= 25, f"Expected at least 25 pages in QA matrix, got {len(PAGES)}"
    assert len(BREAKPOINTS) == 4, "Expected 4 breakpoints (mobile/tablet/laptop/desktop)"
    assert len(BROWSERS) == 3, "Expected 3 browsers (Chromium/WebKit/Firefox)"
    assert len(PAGE_CHECKLIST) >= 8, "Per-page checklist should cover at least 8 items"


def main() -> None:
    assert_structure()

    wb = Workbook()
    build_status_sheet(wb)
    build_test_plan_sheet(wb)
    build_signoff_sheet(wb)
    build_legend_sheet(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    rel = OUTPUT_PATH.relative_to(REPO_ROOT)
    cells = (
        len(PAGES) * (2 + len(BREAKPOINTS) * len(BROWSERS) + 1)
        + len(PAGES) * len(PAGE_CHECKLIST) * 6
    )
    print(f"Wrote {rel} - {cells} cells across 4 sheets.")


if __name__ == "__main__":
    main()
